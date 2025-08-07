import os
import sys
import openpyxl
from openpyxl.utils import column_index_from_string
import argparse
from config import (
    DEFAULT_COLUMN,
    DEFAULT_EXCEL_FILENAME,
    DEFAULT_HEADER,
    DEFAULT_SHEET_NAME,
    INCLUDE_HIDDEN_FILES,
    USE_FULL_PATH
)

def main():
    # Initialize argument parser
    parser = argparse.ArgumentParser(description="Read file names from a given directory, and copy them to an Excel file.")

    # Add arguments for agrument parser
    parser.add_argument(
        "directory",
        type=str,
        help="Path to the directory containing files."
    )
    parser.add_argument(
        "--output", 
        type=str,
        help='Path to the Excel file to write to. If not specified, a new Excel file named "file_names.xlsx" will be created in the directory.',
        default=None
    )
    parser.add_argument(
        "--sheet",
        type=str,
        help='Name of the sheet to write to. Default is the first (active) sheet.',
        default=None
    )
    parser.add_argument(
        "--column",
        type=str,
        help=f'Column letter to insert file names into. Default is "{DEFAULT_COLUMN}".',
        default=DEFAULT_COLUMN
    )
    parser.add_argument(
        "--append",
        action="store_true",
        help="If set, file names will be appended to the existing Excel file (if it exists). Otherwise, it will overwrite or create a new one."
    )

    args = parser.parse_args()

    # Normalize directory path
    directory = os.path.abspath(args.directory)
    if not os.path.isdir(directory):
        print(f"Error: '{directory}' is not a valid directory.")
        sys.exit(1)

    # Set default output path
    output_path = args.output or os.path.join(directory, DEFAULT_EXCEL_FILENAME)

    # Get file names 
    file_names = get_file_list(directory)

    # Prepare workbook
    if args.append and os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()

    # Get target sheet 
    sheet = get_target_sheet(wb, args.sheet or DEFAULT_SHEET_NAME)

    # Determine column index
    try:
        col_index = column_index_from_string(args.column.upper())
    except ValueError:
        print(f"Error: Invlaid column letter '{args.column}'.")
        sys.exit(1)

    # Write header if creating new file or sheet
    start_row = find_first_empty_row(sheet, col_index)
    if start_row == 1:
        sheet.cell(row=1, column=col_index, value=DEFAULT_HEADER)
        start_row += 1

    # Write file names 
    for i, file in enumerate(file_names, start=start_row):
        sheet.cell(row=i, column=col_index, value=file)

    # Save workbook
    wb.save(output_path)
    print(f"File names saved to: {output_path}")


# Help functions 

def get_file_list(directory):
    files = os.listdir(directory)
    if not INCLUDE_HIDDEN_FILES: 
        files = [f for f in files if not f.startswith('.')]
    if USE_FULL_PATH:
        files = [os.path.join(directory, f) for f in files]
    return files

def get_target_sheet(workbook, sheet_name):
    if sheet_name:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        else:
            return workbook.create_sheet(title=sheet_name)
    return workbook.active

def find_first_empty_row(sheet, col_index):
    row = 1
    while sheet.cell(row=row, column=col_index).value:
        row += 1
    return row

if __name__ == "__main__":
    main()